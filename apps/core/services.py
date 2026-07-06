import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse


def export_to_excel(headers, rows, filename='export.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_school_admin_excel(school, admin_login, admin_password):
    return export_to_excel(
        headers=['Школа', 'Район', 'ФИО админа', 'Логин', 'Пароль'],
        rows=[[school.name, school.district.name, f'{school.name} Admin', admin_login, admin_password]],
        filename=f'school_{school.id}_admin.xlsx',
    )


def generate_students_excel(users_with_passwords):
    headers = ['ФИО', 'Класс', 'Логин', 'Пароль']
    rows = [
        [f'{u.last_name} {u.first_name}', str(u.grade) if u.grade else '', u.login, pwd]
        for u, pwd in users_with_passwords
    ]
    return export_to_excel(headers, rows, filename='students.xlsx')


def generate_return_list_excel(students_data, debtors_only=False):
    headers = ['ФИО', 'Кол-во учебников']
    subjects = set()
    for sd in students_data:
        for s in sd.get('subjects', []):
            subjects.add(s['name'])
    subjects = sorted(subjects)
    headers.extend(subjects)
    rows = []
    for sd in students_data:
        if debtors_only and sd.get('all_returned', True):
            continue
        lang = sd.get('language', 'ru')
        row = [sd['fio'], sd['total']]
        for subj in subjects:
            status_map = {s['name']: s['status'] for s in sd.get('subjects', [])}
            val = status_map.get(subj, '')
            returned = {'ru': 'Сдал', 'uz': 'Topshirdi', 'kaa': 'Tapsyrdy', 'en': 'Returned'}
            not_returned = {'ru': 'Не сдал', 'uz': 'Topshirmadi', 'kaa': 'Tapsyrmady', 'en': 'Not returned'}
            row.append(returned.get(lang, 'Сдал') if val == 'returned' else not_returned.get(lang, 'Не сдал') if val == 'active' else '')
        rows.append(row)
    suffix = '_debtors' if debtors_only else ''
    return export_to_excel(headers, rows, filename=f'return_list{suffix}.xlsx')
